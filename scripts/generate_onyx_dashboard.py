
import os
import pandas as pd
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from dotenv import load_dotenv

# Load credentials
load_dotenv()
SUPABASE_URL = os.getenv("VITE_SUPABASE_URL")
SUPABASE_KEY = os.getenv("VITE_SUPABASE_ANON_KEY")

# Vendor Colors from app/src/lib/consts.tsx
VENDOR_COLORS = {
    'R': '737104', 'M': '4F2068', 'W': 'E67E22', 'C': 'D35400',
    'JM': '6BCEBB', 'EM': '00AEEF', 'CA': '85C1E9', 'AN': 'FFED00',
    'SU': 'B19CD9', 'TE': 'FFCB05', 'DH': '8DC63F', 'ML': 'F9A17A',
    'GE': 'F7941D', 'FR': 'F36F21', 'ET': '636466', 'AM': '800020',
    'BT': '603913', 'RF': '00A591', 'GS': 'D11C7E', 'CP': 'A01E5D'
}

STATUS_FILLS = {
    'Paid': PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid'),
    'Requested': PatternFill(start_color='FACC15', end_color='FACC15', fill_type='solid'),
    'Pending': PatternFill(start_color='FACC15', end_color='FACC15', fill_type='solid'),
    'Partial': PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
}

CAT_COLORS = {
    'Acq': '10B981', 'Prod': '6366F1', 'Monthly': '38BDF8', 
    'Oprt': '818CF8', 'Packing': 'FB7185', 'Sppl': '34D399', 'Labr': 'FB7185'
}

def style_header(ws, columns_count):
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    for col in range(1, columns_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def generate_dashboard():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("Error: Supabase credentials missing in .env")
        return

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # 1. Fetch Data
    print("Fetching data from Supabase...")
    finance_res = supabase.table("finance").select("*").execute()
    inventory_res = supabase.table("inventory").select("*").execute()
    
    df_finance = pd.DataFrame(finance_res.data)
    df_inventory = pd.DataFrame(inventory_res.data)
    
    wb = Workbook()
    
    # --- 1. OVERVIEW DASHBOARD ---
    ws_dash = wb.active
    ws_dash.title = "0_Overview"
    ws_dash.append(["ONYX.MX STUDIO - MASTER DASHBOARD", "", "", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_dash.merge_cells('A1:D1')
    ws_dash['A1'].font = Font(size=14, bold=True, color='00AEEF')
    
    ws_dash.append([])
    ws_dash.append(["FINANCIAL SUMMARY", "TOTAL (MXN)", "COUNT"])
    
    total_mxn = df_finance['amount'].sum() + df_finance['commission'].fillna(0).sum()
    ws_dash.append(["Grand Total", total_mxn, len(df_finance)])
    
    paid_group = df_finance.groupby('status')['amount'].sum()
    for status, amt in paid_group.items():
        ws_dash.append([f"Status: {status}", amt])
        
    # --- 2. FINANCE LEDGER ---
    ws_finance = wb.create_sheet("1_Finance_Ledger")
    cols = ['date', 'description', 'category', 'amount', 'commission', 'status', 'vendor_id', 'reference']
    df_f = df_finance[cols].copy()
    
    for r in dataframe_to_rows(df_f, index=False, header=True):
        ws_finance.append(r)
        
    style_header(ws_finance, len(cols))
    
    # Formatting Finance Rows
    for row_idx, row in enumerate(ws_finance.iter_rows(min_row=2), start=2):
        status = ws_finance.cell(row=row_idx, column=6).value
        if status in STATUS_FILLS:
            ws_finance.cell(row=row_idx, column=6).fill = STATUS_FILLS[status]
            ws_finance.cell(row=row_idx, column=6).font = Font(bold=True, color='FFFFFF')
            
        cat = str(ws_finance.cell(row=row_idx, column=3).value)
        for key, color in CAT_COLORS.items():
            if key in cat:
                ws_finance.cell(row=row_idx, column=3).font = Font(color=color, bold=True)

    # --- 3. INVENTORY BY VENDOR ---
    vendors_list = df_inventory['vendor_id'].unique()
    for v_id in vendors_list:
        if not v_id: continue
        ws_v = wb.create_sheet(str(v_id)[:31])
        df_v = df_inventory[df_inventory['vendor_id'] == v_id][['item_id', 'status', 'description', 'price_mxn', 'quantity', 'pay_req']]
        
        for r in dataframe_to_rows(df_v, index=False, header=True):
            ws_v.append(r)
            
        v_color = VENDOR_COLORS.get(v_id, '1F2937')
        header_fill = PatternFill(start_color=v_color, end_color=v_color, fill_type='solid')
        for col in range(1, 7):
            ws_v.cell(row=1, column=col).fill = header_fill
            ws_v.cell(row=1, column=col).font = Font(bold=True, color='FFFFFF')
            
        for row_idx in range(2, ws_v.max_row + 1):
            ws_v.cell(row=row_idx, column=1).font = Font(color=v_color, bold=True)
            
    # Save
    filename = f"Onyx_Master_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Success! Generated: {filename}")

if __name__ == "__main__":
    generate_dashboard()
